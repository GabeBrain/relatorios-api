#!/usr/bin/env python3
"""
fonte_extractor — planilhas de trabalho do analista -> <slug>.fonte.json

Irmão do `ir_extractor.py`: onde aquele extrai o Intermediate Representation do
PPTX, este extrai a VERDADE NUMÉRICA das planilhas que geraram o estudo. O par
alimenta o SOURCE_CROSSCHECK — conferir o deck contra a fonte, uma classe de erro
que nenhuma checagem interna pega (ver FP_sessao_2026-08-12.md, FN-04).

Princípio, medido em 3 pacotes reais (Rolândia, Toledo, Marka): **o nome do
arquivo mente; a aba não.** Os nomes não se repetem entre estudos que cumprem o
mesmo papel, mas as abas são a mesma família. Então:

  1. Ancora na ABA, nunca em coordenada fixa nem em nome de arquivo.
  2. Acha o cabeçalho procurando os NOMES das colunas (a linha varia: 5 em
     jul/2026, 6 em jan/2026).
  3. Normaliza acento, caixa e runs de espaço antes de comparar
     ("Oferta         Final").
  4. Casa por CONCEITO via sinônimos ("oferta atual" == "oferta final").
  5. Falha alto: aba ausente, cabeçalho não encontrado ou papel ambíguo viram
     aviso explícito no JSON, nunca silêncio.
  6. #REF!/#DIV/0! viram célula inválida (null + aviso), jamais zero.
  7. Registra procedência (arquivo, aba, linha) de cada bloco.

Uso:
    python fonte_extractor.py <pasta-excel> --slug <slug> [-o <saida.json>]
    python fonte_extractor.py fontes/housi-toledo-2026-07/excel --slug housi-toledo-2026-07
"""
from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
import unicodedata
import warnings
from datetime import datetime, timezone
from pathlib import Path

# O console do Windows abre em cp1252 e engasga em acento/emoji do relatório.
for _fluxo in (sys.stdout, sys.stderr):
    try:
        _fluxo.reconfigure(encoding="utf-8")
    except Exception:  # pragma: no cover - pipes já em utf-8
        pass

# As planilhas do time usam formatação condicional que o openpyxl não entende;
# o aviso é ruído e não afeta os valores lidos.
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

try:
    import openpyxl
except ImportError:  # pragma: no cover
    sys.exit("openpyxl não instalado: pip install openpyxl")

FONTE_VERSION = 1

# Erros do Excel chegam como string pelo openpyxl(data_only=True). Zero seria
# uma mentira aritmética — viram None + aviso.
ERRO_EXCEL = {"#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?", "#NULL!", "#NUM!"}


def norm(v) -> str:
    """Minúsculo, sem acento, sem pontuação de borda, runs de espaço colapsados."""
    s = unicodedata.normalize("NFD", str(v if v is not None else ""))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", s).strip().lower()


def cell(v):
    """Valor normalizado de célula: número, None, ou string limpa."""
    if v is None:
        return None
    if isinstance(v, str):
        t = v.strip()
        if t in ERRO_EXCEL or t in {"-", "–", "—", ""}:
            return None
        return t
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return v
    return str(v)


def is_erro(v) -> bool:
    return isinstance(v, str) and v.strip() in ERRO_EXCEL


# ---------------------------------------------------------------- papéis

# Classificação por PALAVRA-CHAVE no nome normalizado. Ordem importa: o primeiro
# padrão que casar vence, então o mais específico vem antes.
PAPEIS = [
    # Bases brutas entram no inventário mas não são extraídas: são o insumo, não
    # a saída. Vêm antes para não disputarem papel com a análise correspondente
    # ("00. Base - Revenda e Locação" × "02. Revenda - Apartamento").
    ("base", r"^0*0\.? ?base\b|^base\b"),
    ("oferta", r"consolidada|analise vertical otimizada|esgotados|tabela lacunas"),
    ("socio", r"sociodemografia|onmaps"),
    ("absorcao", r"absorcao|absorçao|cenarios"),
    ("populacao", r"populacao (e|de) domicilios|populacao e domicilios"),
    ("revenda", r"revenda"),
    ("locacao", r"locacao"),
    ("lazer", r"area de lazer"),
    ("anuncios", r"arbnb|airbnb|anuncios"),
    ("pontos_interesse", r"pontos de interesse|polos geradores"),
    ("financiamento", r"financiamento"),
    ("ibge_bruto", r"^dados[_ ]"),
]

# Papéis que são COLEÇÃO por natureza — vários arquivos cumprem o mesmo papel sem
# que isso seja conflito. Alertar aqui seria ruído, e aviso ruidoso é aviso ignorado.
PAPEIS_COLETIVOS = {"base", "ibge_bruto"}

# Recorte territorial/segmento — sai do nome do arquivo, é o único uso legítimo dele.
RECORTES = [
    (r"\b1\s*km\b", "1 km"), (r"\b2\s*km\b", "2 km"), (r"\b3\s*km\b", "3 km"),
    (r"z\.?i\.?\s*total", "z.i. total"), (r"\bz\.?i\.?\b", "z.i."),
    (r"his\s*1", "his 1"), (r"his\s*2", "his 2"), (r"\bhmp\b", "hmp"),
    (r"primaria", "primaria"), (r"\bcidade\b", "cidade"),
]


def classifica(nome: str):
    # `_` e `-` viram espaço: "1 Km_verificação" precisa casar com \b1 km\b, e
    # \b não fecha antes de underscore (é caractere de palavra em regex).
    n = re.sub(r"[_\-]+", " ", norm(nome))
    n = re.sub(r"\s+", " ", n)
    papel = next((p for p, rx in PAPEIS if re.search(rx, n)), None)
    recorte = next((r for rx, r in RECORTES if re.search(rx, n)), None)
    return papel, recorte


# ---------------------------------------------------------------- conceitos

# Sinônimos por conceito. Medidos: jul/2026 usa "Oferta Atual"/"Vendas";
# jan/2026 usa "Oferta Final"/"Vendas s/ O.L.". Mesmo conceito, rótulo diferente.
CONCEITOS = {
    "n_empreendimentos": [r"^n[ºo°]? de empreend"],
    "oferta_lancada": [r"^oferta lancada$"],
    "oferta_atual": [r"^oferta atual$", r"^oferta final$"],
    "vendas": [r"^vendas$", r"^vendas s/ ?o\.?l\.?$"],
    "disponibilidade": [r"^disp\.? sobre lancados$", r"^disp\.? s/ ?o\.?l\.?$"],
}

# Abas de saída por papel, com a coluna-âncora que identifica o cabeçalho.
BLOCOS_OFERTA = {
    "padrao": ("Padrão", "oferta lancada"),
    "ano": ("Ano", "oferta lancada"),
    "tipologia": ("Tipologia", "oferta lancada"),
}
# Sociodemografia: cabeçalho de DOIS níveis (recortes em cima, Absoluto/% embaixo).
BLOCOS_SOCIO = {
    "domicilios_por_tipo": "Dom.p Tipo",
    "condicao_ocupacao": "Dom.p Cond. Ocup.",
    "populacao_faixa_etaria": "População - Faixa Etária",
    "domicilios_por_moradores": "Dom.p nº Moradores",
}


def acha_aba(wb, alvo: str):
    """Aba por nome normalizado — tolera acento, caixa e espaço."""
    a = norm(alvo)
    for título in wb.sheetnames:
        if norm(título) == a:
            return wb[título]
    for título in wb.sheetnames:  # fallback: prefixo (ex.: "Dom.p Tipo (2)")
        if norm(título).startswith(a):
            return wb[título]
    return None


def linhas_da(ws, limite=400):
    return [list(r) for r in ws.iter_rows(values_only=True)][:limite]


def acha_cabecalho(linhas, ancora: str, ate=20):
    """Índice (0-based) da linha cujo conteúdo contém a coluna-âncora."""
    for i, row in enumerate(linhas[:ate]):
        if any(norm(c) == ancora for c in row if c is not None):
            return i
    return None


def mapeia_conceitos(row):
    """coluna -> conceito, para as colunas que reconhecemos."""
    out = {}
    for j, c in enumerate(row):
        n = norm(c)
        if not n:
            continue
        for conceito, padroes in CONCEITOS.items():
            if any(re.match(p, n) for p in padroes):
                out[j] = conceito
                break
    return out


def extrai_oferta(wb, arquivo, papel_rec, avisos):
    """Padrão / Ano / Tipologia — uma linha por categoria + linha Total."""
    blocos = []
    for tabela, (aba_nome, ancora) in BLOCOS_OFERTA.items():
        ws = acha_aba(wb, aba_nome)
        if ws is None:
            avisos.append({"tipo": "aba_ausente", "arquivo": arquivo, "aba": aba_nome})
            continue
        linhas = linhas_da(ws)
        h = acha_cabecalho(linhas, ancora)
        if h is None:
            avisos.append({"tipo": "cabecalho_nao_encontrado", "arquivo": arquivo,
                           "aba": aba_nome, "ancora": ancora})
            continue
        colunas = mapeia_conceitos(linhas[h])
        if not colunas:
            avisos.append({"tipo": "nenhuma_coluna_reconhecida", "arquivo": arquivo, "aba": aba_nome})
            continue
        # rótulo = primeira coluna textual do cabeçalho (Padrão / Ano Lançamento / Tipologia)
        col_rotulo = next((j for j, c in enumerate(linhas[h]) if norm(c) and j not in colunas), 0)

        itens, total = [], None
        for i in range(h + 1, len(linhas)):
            row = linhas[i]
            rotulo = cell(row[col_rotulo]) if col_rotulo < len(row) else None
            valores, invalidas = {}, []
            for j, conceito in colunas.items():
                bruto = row[j] if j < len(row) else None
                if is_erro(bruto):
                    invalidas.append(conceito)
                valores[conceito] = cell(bruto)
            if invalidas:
                avisos.append({"tipo": "celula_invalida", "arquivo": arquivo, "aba": aba_nome,
                               "linha": i + 1, "conceitos": invalidas})
            if rotulo is None and not any(v is not None for v in valores.values()):
                continue
            registro = {"rotulo": rotulo, "linha": i + 1, "valores": valores}
            if rotulo and norm(rotulo).startswith("total"):
                total = registro
                break
            itens.append(registro)

        if total is None:
            avisos.append({"tipo": "sem_linha_total", "arquivo": arquivo, "aba": aba_nome})
        blocos.append({
            "papel": "oferta", "tabela": tabela, **papel_rec,
            "arquivo": arquivo, "aba": ws.title, "cabecalho_linha": h + 1,
            "conceitos": sorted(set(colunas.values())),
            "itens": itens, "total": total,
        })
    return blocos


def extrai_socio(wb, arquivo, papel_rec, avisos):
    """
    Cabeçalho de dois níveis: linha com os recortes (PR | cidade | Até 1 km …) e,
    abaixo, a linha Absoluto/%. Os recortes vêm de células mescladas, então o
    rótulo aparece uma vez e as seguintes ficam vazias — preenchemos à frente.
    """
    blocos = []
    for tabela, aba_nome in BLOCOS_SOCIO.items():
        ws = acha_aba(wb, aba_nome)
        if ws is None:
            avisos.append({"tipo": "aba_ausente", "arquivo": arquivo, "aba": aba_nome})
            continue
        linhas = linhas_da(ws, limite=80)
        h = next((i for i, row in enumerate(linhas[:15])
                  if sum(1 for c in row if norm(c) == "absoluto") >= 2), None)
        if h is None or h == 0:
            avisos.append({"tipo": "cabecalho_nao_encontrado", "arquivo": arquivo,
                           "aba": aba_nome, "ancora": "absoluto"})
            continue

        # Recortes na linha de cima (células mescladas: o rótulo aparece uma vez).
        # Cada recorte é dono de EXATAMENTE um "Absoluto" e um "%" — os primeiros
        # a partir da sua coluna. Propagar rótulo à frente sem esse teto faz as
        # colunas órfãs do fim da planilha (restos de fórmula quebrada) herdarem
        # o último recorte e sobrescreverem os valores certos — foi o que
        # aconteceu com "Até 3 Km" na Rolândia, que virou zero.
        cabec = linhas[h]
        tipo_de = lambda j: norm(cabec[j]) if j < len(cabec) else ""
        # A primeira coluna "Absoluto" abre a área de dados. O que estiver à
        # esquerda dela é o cabeçalho da coluna de rótulos ("TIPO DE DOMICÍLIO",
        # "MORADORES"), não um recorte.
        inicio = next((k for k in range(len(cabec)) if tipo_de(k) == "absoluto"), 0)
        rotulos = [(j, str(cell(c)).strip()) for j, c in enumerate(linhas[h - 1])
                   if cell(c) and j >= inicio]
        recortes, tipos = {}, {}
        for idx, (col, rot) in enumerate(rotulos):
            fim = rotulos[idx + 1][0] if idx + 1 < len(rotulos) else len(cabec)
            for alvo in ("absoluto", "%"):
                j = next((k for k in range(col, fim) if tipo_de(k) == alvo), None)
                if j is not None:
                    recortes[j], tipos[j] = rot, alvo
            if not any(v == rot for v in recortes.values()):
                avisos.append({"tipo": "recorte_sem_coluna", "arquivo": arquivo,
                               "aba": aba_nome, "recorte": rot})

        itens, total = [], None
        for i in range(h + 1, len(linhas)):
            row = linhas[i]
            rotulo = next((cell(c) for c in row[:3] if isinstance(cell(c), str)), None)
            if not rotulo or norm(rotulo).startswith("fonte"):
                if total is not None:
                    break
                continue
            por_recorte, invalidas = {}, []
            for j, rec in recortes.items():
                bruto = row[j] if j < len(row) else None
                if is_erro(bruto):
                    invalidas.append(rec)
                v = cell(bruto)
                if v is None:
                    continue
                por_recorte.setdefault(rec, {})[tipos[j]] = v
            if invalidas:
                avisos.append({"tipo": "celula_invalida", "arquivo": arquivo, "aba": aba_nome,
                               "linha": i + 1, "recortes": invalidas})
            if not por_recorte:
                continue
            registro = {"rotulo": str(rotulo).strip(), "linha": i + 1, "recortes": por_recorte}
            if norm(rotulo).startswith("total"):
                total = registro
                break
            itens.append(registro)

        blocos.append({
            "papel": "socio", "tabela": tabela, **papel_rec,
            "arquivo": arquivo, "aba": ws.title, "cabecalho_linha": h + 1,
            "recortes": sorted(set(recortes.values())),
            "itens": itens, "total": total,
        })
    return blocos


def extrai_taxas(wb, arquivo, papel_rec, avisos):
    """TAXAS da absorção — as constantes que o deck cita em texto ('X% a.a.')."""
    ws = acha_aba(wb, "TAXAS")
    if ws is None:
        avisos.append({"tipo": "aba_ausente", "arquivo": arquivo, "aba": "TAXAS"})
        return []
    linhas = linhas_da(ws, limite=40)
    achadas = []
    for i, row in enumerate(linhas):
        titulo = next((cell(c) for c in row if isinstance(cell(c), str)), None)
        for j, c in enumerate(row):
            v = cell(c)
            # taxa anual plausível: fração entre 0 e 0.5, com rótulo à esquerda
            if isinstance(v, (int, float)) and 0 < v < 0.5:
                achadas.append({"linha": i + 1, "coluna": j + 1, "valor": round(v, 6),
                                "contexto": str(titulo)[:60] if titulo else None})
    return [{"papel": "absorcao", "tabela": "taxas", **papel_rec,
             "arquivo": arquivo, "aba": ws.title, "valores": achadas}]


def extrai_populacao(wb, arquivo, papel_rec, avisos):
    """População / Domicilios — série 2000/2010/2026 + taxa de variação anual."""
    blocos = []
    for tabela, aba_nome in (("populacao", "População"), ("domicilios", "Domicilios")):
        ws = acha_aba(wb, aba_nome)
        if ws is None:
            avisos.append({"tipo": "aba_ausente", "arquivo": arquivo, "aba": aba_nome})
            continue
        linhas = linhas_da(ws, limite=40)
        itens = []
        for i, row in enumerate(linhas):
            vals = [cell(c) for c in row]
            rotulo = next((v for v in vals[:3] if isinstance(v, str)), None)
            numeros = [v for v in vals if isinstance(v, (int, float))]
            if rotulo and len(numeros) >= 3 and not norm(rotulo).startswith("fonte"):
                itens.append({"rotulo": str(rotulo).strip(), "linha": i + 1, "numeros": numeros[:8]})
        blocos.append({"papel": "populacao", "tabela": tabela, **papel_rec,
                       "arquivo": arquivo, "aba": ws.title, "itens": itens})
    return blocos


EXTRATORES = {
    "oferta": extrai_oferta,
    "socio": extrai_socio,
    "absorcao": extrai_taxas,
    "populacao": extrai_populacao,
}


def main() -> int:
    ap = argparse.ArgumentParser(description="Extrai a verdade numérica das planilhas do analista.")
    ap.add_argument("pasta", help="pasta com os .xlsx/.xlsm do estudo")
    ap.add_argument("--slug", required=True, help="identificador do estudo (ex.: housi-toledo-2026-07)")
    ap.add_argument("-o", "--saida", help="arquivo .fonte.json (padrão: <slug>.fonte.json na pasta atual)")
    args = ap.parse_args()

    pasta = Path(args.pasta)
    if not pasta.is_dir():
        return print(f"✖ pasta inexistente: {pasta}") or 2

    arquivos = sorted(p for p in pasta.iterdir() if p.suffix.lower() in {".xlsx", ".xlsm"})
    if not arquivos:
        return print(f"✖ nenhuma planilha em {pasta}") or 2

    avisos, blocos, inventario = [], [], []
    vistos = {}

    for caminho in arquivos:
        papel, recorte = classifica(caminho.name)
        sha1 = hashlib.sha1(caminho.read_bytes()).hexdigest()
        inventario.append({"arquivo": caminho.name, "sha1": sha1[:16],
                           "papel": papel, "recorte": recorte,
                           "tamanho_kb": round(caminho.stat().st_size / 1024, 1)})
        if papel is None:
            avisos.append({"tipo": "papel_desconhecido", "arquivo": caminho.name})
            continue
        # Dois arquivos disputando papel+recorte é ambiguidade REAL (as duas
        # ABSORCAO do Toledo: Z.I. x cidade). Registrar, nunca escolher calado.
        chave = (papel, recorte)
        if chave in vistos and papel not in PAPEIS_COLETIVOS:
            avisos.append({"tipo": "papel_ambiguo", "papel": papel, "recorte": recorte,
                           "arquivos": [vistos[chave], caminho.name],
                           "acao": "ambos extraidos; desempate manual necessario"})
        vistos.setdefault(chave, caminho.name)

        extrator = EXTRATORES.get(papel)
        if extrator is None:
            continue
        try:
            wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
        except Exception as exc:  # planilha corrompida não pode derrubar o pacote
            avisos.append({"tipo": "falha_ao_abrir", "arquivo": caminho.name, "erro": str(exc)[:120]})
            continue
        try:
            blocos.extend(extrator(wb, caminho.name, {"recorte": recorte}, avisos))
        finally:
            wb.close()

    saida = {
        "fonte_version": FONTE_VERSION,
        "estudo": args.slug,
        "gerado_em": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "pasta": str(pasta).replace("\\", "/"),
        "inventario": inventario,
        "blocos": blocos,
        "avisos": avisos,
    }
    destino = Path(args.saida) if args.saida else Path(f"{args.slug}.fonte.json")
    destino.parent.mkdir(parents=True, exist_ok=True)
    destino.write_text(json.dumps(saida, ensure_ascii=False, indent=2), encoding="utf-8")

    por_tipo = {}
    for a in avisos:
        por_tipo[a["tipo"]] = por_tipo.get(a["tipo"], 0) + 1
    print(f"✅ {destino}  ({destino.stat().st_size / 1024:.1f} KB)")
    print(f"   {len(arquivos)} planilhas · {len(blocos)} blocos extraídos")
    for papel in sorted({b['papel'] for b in blocos}):
        n = sum(1 for b in blocos if b["papel"] == papel)
        print(f"      {papel}: {n} bloco(s)")
    if por_tipo:
        print(f"   avisos: " + " · ".join(f"{k}={v}" for k, v in sorted(por_tipo.items())))
    return 0


if __name__ == "__main__":
    sys.exit(main())
